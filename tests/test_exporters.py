"""Phase 6 tests: the four exporters.

Every exporter is checked for protocol conformance, deterministic output
(export twice, compare), input immutability, and correct content. The
golden-example test runs the full six-stage pipeline once and exports the
real result through all four formats. No exporter makes an LLM call."""

import csv
import io
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from qaops.config import QAOpsSettings
from qaops.core.protocols import Exporter
from qaops.exporters import CsvExporter, ExcelExporter, JsonExporter, MarkdownExporter
from qaops.llm import MockLLMClient, PromptLoader
from qaops.models import (
    BusinessRule,
    Requirement,
    RequirementInput,
    Scenario,
    ScenarioCategory,
    TestCase,
    TestDesignResult,
    TestStep,
)
from qaops.pipelines.test_design import CoverageValidator, build_full_pipeline

EXAMPLES_DIR = Path(__file__).resolve().parent.parent / "examples"


def sample_result() -> TestDesignResult:
    base = TestDesignResult(
        source_name="login.md",
        requirements=[
            Requirement(id="REQ-001", title="Login", description="User logs in.", actors=["User"]),
            Requirement(id="REQ-002", title="Lockout", description="Locks after 5 attempts."),
        ],
        business_rules=[
            BusinessRule(id="BR-001", requirement_id="REQ-002", rule="Locks after 5 attempts."),
        ],
        scenarios=[
            Scenario(
                id="SC-001",
                title="valid login",
                category=ScenarioCategory.POSITIVE,
                requirement_ids=["REQ-001"],
            ),
            Scenario(
                id="SC-002",
                title="lockout at fifth attempt",
                category=ScenarioCategory.BOUNDARY_VALUE,
                requirement_ids=["REQ-002"],
            ),
        ],
        test_cases=[
            TestCase(
                id="TC-001",
                scenario_id="SC-001",
                requirement_ids=["REQ-001"],
                module="Auth",
                feature="Login",
                title="Valid login redirects to dashboard",
                objective="Prove a valid login works",
                preconditions=["A registered user exists"],
                test_data={"email": "user@example.com", "password": "Valid@123"},
                steps=[
                    TestStep(number=1, action="Open login page", expected="Form shown"),
                    TestStep(number=2, action="Submit valid credentials"),
                ],
                expected_result="User lands on dashboard",
                tags=["smoke", "login"],
            ),
        ],
    )
    return CoverageValidator().run(base)


ALL_EXPORTERS = [JsonExporter, MarkdownExporter, CsvExporter, ExcelExporter]


class TestProtocolConformance:
    @pytest.mark.parametrize("exporter_cls", ALL_EXPORTERS)
    def test_satisfies_protocol(self, exporter_cls: type) -> None:
        exporter = exporter_cls()
        assert isinstance(exporter, Exporter)
        assert exporter.format_name
        assert exporter.file_extension.startswith(".")


class TestImmutability:
    @pytest.mark.parametrize("exporter_cls", ALL_EXPORTERS)
    def test_export_does_not_mutate_input(self, exporter_cls: type, tmp_path: Path) -> None:
        result = sample_result()
        before = result.model_dump_json()
        path = tmp_path / f"out{exporter_cls().file_extension}"
        exporter_cls().export(result, str(path))
        assert result.model_dump_json() == before


class TestDeterminism:
    def test_json_byte_identical_across_runs(self, tmp_path: Path) -> None:
        result = sample_result()
        a = tmp_path / "a.json"
        b = tmp_path / "b.json"
        JsonExporter().export(result, str(a))
        JsonExporter().export(result, str(b))
        assert a.read_bytes() == b.read_bytes()

    def test_markdown_identical_across_runs(self) -> None:
        result = sample_result()
        exporter = MarkdownExporter()
        assert exporter.to_markdown(result) == exporter.to_markdown(result)

    def test_csv_identical_across_runs(self) -> None:
        result = sample_result()
        exporter = CsvExporter()
        assert exporter.to_csv(result) == exporter.to_csv(result)

    def test_excel_cell_values_identical_across_runs(self, tmp_path: Path) -> None:
        result = sample_result()
        a = tmp_path / "a.xlsx"
        b = tmp_path / "b.xlsx"
        ExcelExporter().export(result, str(a))
        ExcelExporter().export(result, str(b))
        cells_a = _all_cell_values(a)
        cells_b = _all_cell_values(b)
        assert cells_a == cells_b

    def test_no_timestamp_in_json(self) -> None:
        text = JsonExporter().to_json(sample_result())
        lowered = text.lower()
        for banned in ("timestamp", "generated_at", "created_at", "2026-", "datetime"):
            assert banned not in lowered


class TestJsonExporter:
    def test_round_trips_back_into_model(self) -> None:
        result = sample_result()
        text = JsonExporter().to_json(result)
        restored = TestDesignResult.model_validate_json(text)
        assert restored == result

    def test_file_ends_with_newline(self, tmp_path: Path) -> None:
        path = tmp_path / "out.json"
        JsonExporter().export(result=sample_result(), output_path=str(path))
        assert path.read_text(encoding="utf-8").endswith("}\n")


class TestMarkdownExporter:
    def test_contains_all_sections(self) -> None:
        md = MarkdownExporter().to_markdown(sample_result())
        assert "# Test Design Report: login.md" in md
        assert "## Coverage Summary" in md
        assert "## Requirements" in md
        assert "## Scenarios" in md
        assert "## Test Cases" in md
        assert "TC-001" in md and "REQ-001" in md and "SC-001" in md

    def test_coverage_percentages_rendered(self) -> None:
        md = MarkdownExporter().to_markdown(sample_result())
        assert "50.0%" in md  # 1 of 2 requirements covered


class TestCsvExporter:
    def test_one_row_per_test_case_with_stable_columns(self) -> None:
        text = CsvExporter().to_csv(sample_result())
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
        assert len(rows) == 1
        row = rows[0]
        assert row["id"] == "TC-001"
        assert row["requirement_ids"] == "REQ-001"
        assert row["tags"] == "smoke; login"
        assert "1. Open login page -> Form shown" in row["steps"]
        assert reader.fieldnames is not None and reader.fieldnames[0] == "id"


class TestExcelExporter:
    def test_workbook_has_expected_sheets(self, tmp_path: Path) -> None:
        path = tmp_path / "out.xlsx"
        ExcelExporter().export(sample_result(), str(path))
        wb = load_workbook(path)
        assert wb.sheetnames == [
            "Coverage Summary",
            "Requirements",
            "Scenarios",
            "Test Cases",
            "Traceability",
        ]

    def test_test_cases_sheet_content(self, tmp_path: Path) -> None:
        path = tmp_path / "out.xlsx"
        ExcelExporter().export(sample_result(), str(path))
        sheet = load_workbook(path)["Test Cases"]
        assert sheet["A1"].value == "ID"
        assert sheet["A2"].value == "TC-001"
        assert sheet["A1"].font.name == "Arial"


def _all_cell_values(path: Path) -> dict[str, list[list[object]]]:
    wb = load_workbook(path)
    return {
        name: [[c.value for c in row] for row in wb[name].iter_rows()] for name in wb.sheetnames
    }


# --- Golden examples through the full pipeline then all four exporters --------

ANALYZER_RESPONSE = json.dumps(
    {
        "requirements": [
            {"title": "Primary", "description": "Main behavior.", "actors": ["User"]},
            {"title": "Secondary", "description": "Second behavior."},
        ]
    }
)
RULES_RESPONSE = json.dumps(
    {"rules": [{"requirement_id": "REQ-001", "rule": "a rule", "source_excerpt": ""}]}
)
GAPS_RESPONSE = json.dumps({"gaps": []})
SCENARIOS_RESPONSE = json.dumps(
    {
        "scenarios": [
            {
                "title": "Primary happy path",
                "description": "d",
                "category": "positive",
                "requirement_ids": ["REQ-001"],
            },
            {
                "title": "Secondary boundary",
                "description": "d",
                "category": "boundary_value",
                "requirement_ids": ["REQ-002"],
            },
        ]
    }
)
TEST_CASES_RESPONSE = json.dumps(
    {
        "test_cases": [
            {
                "scenario_id": "SC-001",
                "requirement_ids": ["REQ-001"],
                "title": "Primary case",
                "expected_result": "works",
                "steps": [{"action": "do it", "expected": "done"}],
                "priority": "high",
                "test_type": "functional",
            },
            {
                "scenario_id": "SC-002",
                "requirement_ids": ["REQ-002"],
                "title": "Boundary case",
                "expected_result": "handled",
                "steps": [{"action": "push limit", "expected": "rejected"}],
                "priority": "critical",
                "test_type": "boundary",
            },
        ]
    }
)


@pytest.fixture
def settings(tmp_path: Path) -> QAOpsSettings:
    return QAOpsSettings(output_dir=tmp_path / "out")


@pytest.fixture
def prompts() -> PromptLoader:
    return PromptLoader()


class TestGoldenExampleExport:
    @pytest.mark.parametrize(
        "example", ["login.md", "checkout.md", "video_playback.md", "fund_transfer.md"]
    )
    def test_all_formats_export_from_full_pipeline(
        self, settings: QAOpsSettings, prompts: PromptLoader, example: str, tmp_path: Path
    ) -> None:
        text = (EXAMPLES_DIR / example).read_text(encoding="utf-8")
        mock = MockLLMClient(
            [
                ANALYZER_RESPONSE,
                RULES_RESPONSE,
                GAPS_RESPONSE,
                SCENARIOS_RESPONSE,
                TEST_CASES_RESPONSE,
            ]
        )
        result = build_full_pipeline(mock, prompts, settings).run(
            RequirementInput(text=text, source_name=example)
        )
        assert isinstance(result, TestDesignResult)
        assert mock.call_count == 5  # exporters add no LLM call

        stem = tmp_path / Path(example).stem
        json_path = JsonExporter().export(result, f"{stem}.json")
        md_path = MarkdownExporter().export(result, f"{stem}.md")
        csv_path = CsvExporter().export(result, f"{stem}.csv")
        xlsx_path = ExcelExporter().export(result, f"{stem}.xlsx")

        # JSON round-trips; every artifact is non-empty and readable.
        assert TestDesignResult.model_validate_json(Path(json_path).read_text()) == result
        assert Path(md_path).stat().st_size > 0
        assert len(list(csv.DictReader(io.StringIO(Path(csv_path).read_text())))) == 2
        assert load_workbook(xlsx_path).sheetnames[0] == "Coverage Summary"

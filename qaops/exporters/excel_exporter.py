"""ExcelExporter - a multi-sheet workbook (ADR-016).

Unlike CSV, Excel carries the whole graph across sheets: Coverage
Summary, Requirements, Scenarios, Test Cases, and Traceability. Values
only - no formulas - so the file is deterministic and needs no
recalculation. A professional font is applied per the spreadsheet
conventions.

openpyxl ships as the optional 'excel' extra; import errors surface a
clear message pointing at `pip install qaops-ai[excel]`.

Note on determinism: openpyxl writes a fixed workbook structure for
identical input. It does embed no timestamps in the cells we write; the
zip container's internal metadata is not part of the logical content and
is not asserted on. Tests compare the read-back cell values, which are
fully deterministic.
"""

from typing import TYPE_CHECKING, Any

from qaops.core.errors import ExportError
from qaops.exporters._base import format_steps, join_list, join_test_data, to_canonical_dict
from qaops.models import TestDesignResult

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

_FONT_NAME = "Arial"


class ExcelExporter:
    """Serializes a TestDesignResult to a multi-sheet .xlsx workbook."""

    @property
    def format_name(self) -> str:
        return "xlsx"

    @property
    def file_extension(self) -> str:
        return ".xlsx"

    def export(self, result: TestDesignResult, output_path: str) -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError as exc:  # pragma: no cover - exercised via message only
            msg = (
                "Excel export requires openpyxl. Install the optional extra: "
                "pip install 'qaops-ai[excel]'."
            )
            raise ExportError(msg) from exc

        data = to_canonical_dict(result)
        workbook = Workbook()
        workbook.remove(workbook.active)  # drop the default empty sheet

        header_font = Font(name=_FONT_NAME, bold=True)
        body_font = Font(name=_FONT_NAME)

        self._coverage_sheet(workbook, data, header_font, body_font)
        self._requirements_sheet(workbook, data, header_font, body_font)
        self._scenarios_sheet(workbook, data, header_font, body_font)
        self._test_cases_sheet(workbook, data, header_font, body_font)
        self._traceability_sheet(workbook, data, header_font, body_font)

        workbook.save(output_path)
        return output_path

    # --- sheets --------------------------------------------------------------

    def _write_table(
        self,
        sheet: "Worksheet",
        headers: list[str],
        rows: list[list[Any]],
        header_font: Any,
        body_font: Any,
    ) -> None:
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = header_font
        for row in rows:
            sheet.append(row)
        for row_cells in sheet.iter_rows(min_row=2):
            for cell in row_cells:
                cell.font = body_font

    def _coverage_sheet(self, workbook: Any, data: dict[str, Any], hf: Any, bf: Any) -> None:
        sheet = workbook.create_sheet("Coverage Summary")
        m = data["coverage"]["metrics"]
        rows = [
            ["Requirements", m["covered_requirements"], m["total_requirements"]],
            ["Business rules", m["covered_business_rules"], m["total_business_rules"]],
            ["Scenarios", m["covered_scenarios"], m["total_scenarios"]],
            ["Test cases", "", m["total_test_cases"]],
        ]
        self._write_table(sheet, ["Metric", "Covered", "Total"], rows, hf, bf)

    def _requirements_sheet(self, workbook: Any, data: dict[str, Any], hf: Any, bf: Any) -> None:
        sheet = workbook.create_sheet("Requirements")
        status_by_id = {
            rc["requirement_id"]: rc["status"] for rc in data["coverage"]["per_requirement"]
        }
        rows = [
            [
                req["id"],
                req["title"],
                req["description"],
                join_list(req["actors"]),
                status_by_id.get(req["id"], ""),
            ]
            for req in data["requirements"]
        ]
        self._write_table(sheet, ["ID", "Title", "Description", "Actors", "Coverage"], rows, hf, bf)

    def _scenarios_sheet(self, workbook: Any, data: dict[str, Any], hf: Any, bf: Any) -> None:
        sheet = workbook.create_sheet("Scenarios")
        status_by_id = {sc["scenario_id"]: sc["status"] for sc in data["coverage"]["per_scenario"]}
        rows = [
            [
                sc["id"],
                sc["title"],
                sc["category"],
                join_list(sc["requirement_ids"]),
                status_by_id.get(sc["id"], ""),
            ]
            for sc in data["scenarios"]
        ]
        self._write_table(
            sheet, ["ID", "Title", "Category", "Requirements", "Coverage"], rows, hf, bf
        )

    def _test_cases_sheet(self, workbook: Any, data: dict[str, Any], hf: Any, bf: Any) -> None:
        sheet = workbook.create_sheet("Test Cases")
        headers = [
            "ID",
            "Title",
            "Scenario",
            "Requirements",
            "Module",
            "Feature",
            "Priority",
            "Type",
            "Objective",
            "Preconditions",
            "Test Data",
            "Steps",
            "Expected Result",
        ]
        rows = [
            [
                tc["id"],
                tc["title"],
                tc["scenario_id"],
                join_list(tc["requirement_ids"]),
                tc["module"],
                tc["feature"],
                tc["priority"],
                tc["test_type"],
                tc["objective"],
                join_list(tc["preconditions"]),
                join_test_data(tc["test_data"]),
                format_steps(tc["steps"]),
                tc["expected_result"],
            ]
            for tc in data["test_cases"]
        ]
        self._write_table(sheet, headers, rows, hf, bf)

    def _traceability_sheet(self, workbook: Any, data: dict[str, Any], hf: Any, bf: Any) -> None:
        sheet = workbook.create_sheet("Traceability")
        entries = data["coverage"]["traceability"]["entries"]
        rows = [[req_id, join_list(tc_ids)] for req_id, tc_ids in entries.items()]
        self._write_table(sheet, ["Requirement", "Test Cases"], rows, hf, bf)

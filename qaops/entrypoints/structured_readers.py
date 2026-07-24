"""Structured readers for human-authored scenario documents (ADR-024).

QA teams keep scenarios in spreadsheets and markdown tables, not in QAOps CSV
exports. These readers turn those artifacts into the same record dicts the
existing CSV and JSON paths produce, so `parse_scenarios` treats every format
identically and the pipeline sees only canonical domain models.

They are strictly deterministic - no LLM, no inference. Each reader recognises
an explicit structure (spreadsheet rows, a markdown table, a bulleted or
numbered list) and fails with a clear message when the document is
unstructured prose, rather than guessing. Prose belongs to the `document`
entry point, where the requirement analyzer handles it with a model.

Column and field names are matched case-insensitively and tolerate spaces or
underscores, so 'Requirement IDs', 'requirement_ids', and 'REQUIREMENT IDS'
are the same column.
"""

import re
from pathlib import Path
from typing import Any

from qaops.core.errors import DocumentLoadError
from qaops.models import ScenarioCategory

_KNOWN_CATEGORIES = frozenset(c.value for c in ScenarioCategory)

# Canonical field names, and the aliases a human-authored document may use.
_FIELD_ALIASES: dict[str, str] = {
    "title": "title",
    "scenario": "title",
    "scenarioname": "title",
    "name": "title",
    "summary": "title",
    "description": "description",
    "details": "description",
    "detail": "description",
    "notes": "description",
    "category": "category",
    "type": "category",
    "scenariotype": "category",
    "requirementids": "requirement_ids",
    "requirementid": "requirement_ids",
    "requirements": "requirement_ids",
    "requirement": "requirement_ids",
    "reqids": "requirement_ids",
    "reqid": "requirement_ids",
    "traceability": "requirement_ids",
}

_MARKDOWN_TABLE_SEPARATOR = re.compile(r"^\s*\|?[\s:|-]+\|[\s:|-]*$")
_LIST_ITEM = re.compile(r"^\s*(?:[-*+]|\d+[.)])\s+(.*)$")
_ID_PATTERN = re.compile(r"\b[A-Z]{2,}-\d+\b")


def _normalize_key(raw: str) -> str:
    """Fold a header into a canonical field name, or return it lowercased."""
    key = re.sub(r"[\s_\-]+", "", str(raw).strip().casefold())
    return _FIELD_ALIASES.get(key, key)


def _split_ids(value: str) -> list[str]:
    """Split a requirement-reference cell on any common separator."""
    if not value or not str(value).strip():
        return []
    parts = re.split(r"[;,/|]+", str(value))
    return [p.strip() for p in parts if p.strip()]


def _record_from_mapping(mapping: dict[str, Any]) -> dict[str, Any]:
    """Build a scenario record from an already-keyed mapping."""
    normalized = {_normalize_key(k): v for k, v in mapping.items()}
    return {
        "title": str(normalized.get("title", "") or "").strip(),
        "description": str(normalized.get("description", "") or "").strip(),
        "category": str(normalized.get("category", "") or "").strip() or "functional",
        "requirement_ids": _split_ids(str(normalized.get("requirement_ids", "") or "")),
    }


# --- XLSX --------------------------------------------------------------------


def read_xlsx_scenarios(path: Path) -> list[dict[str, Any]]:
    """Read scenarios from the first worksheet of an .xlsx file.

    The first non-empty row is treated as the header. Requires a recognisable
    title column; other columns are optional.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        msg = (
            "Reading .xlsx input requires openpyxl. Install it with: pip install 'qaops-ai[excel]'."
        )
        raise DocumentLoadError(msg) from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (OSError, ValueError, KeyError) as exc:
        msg = f"Could not read spreadsheet {path}: {exc}"
        raise DocumentLoadError(msg) from exc

    try:
        sheet = workbook.worksheets[0]
        rows = [
            ["" if cell is None else str(cell).strip() for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
    finally:
        workbook.close()

    rows = [row for row in rows if any(cell for cell in row)]
    if not rows:
        msg = f"{path} contains no data."
        raise DocumentLoadError(msg)

    header = [_normalize_key(cell) for cell in rows[0]]
    if "title" not in header:
        msg = (
            f"{path} has no recognisable scenario title column. "
            f"Found headers: {rows[0]}. Expected one of: title, scenario, name, summary."
        )
        raise DocumentLoadError(msg)

    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        mapping = {header[i]: row[i] for i in range(min(len(header), len(row)))}
        record = _record_from_mapping(mapping)
        if record["title"]:
            records.append(record)
    return records


# --- Markdown ----------------------------------------------------------------


def _read_markdown_table(lines: list[str]) -> list[dict[str, Any]]:
    """Parse the first markdown pipe table found, if any."""
    for index, line in enumerate(lines):
        if "|" not in line or index + 1 >= len(lines):
            continue
        if not _MARKDOWN_TABLE_SEPARATOR.match(lines[index + 1]):
            continue
        header = [_normalize_key(c) for c in line.strip().strip("|").split("|")]
        if "title" not in header:
            continue
        records: list[dict[str, Any]] = []
        for row_line in lines[index + 2 :]:
            if "|" not in row_line:
                break
            cells = [c.strip() for c in row_line.strip().strip("|").split("|")]
            mapping = {header[i]: cells[i] for i in range(min(len(header), len(cells)))}
            record = _record_from_mapping(mapping)
            if record["title"]:
                records.append(record)
        return records
    return []


def _read_list_items(lines: list[str]) -> list[dict[str, Any]]:
    """Parse a bulleted or numbered list, one scenario per item.

    Requirement IDs are picked up from any REQ-001 style token in the item, and
    a trailing parenthesised word is read as the category when it matches one.
    """
    records: list[dict[str, Any]] = []
    for line in lines:
        match = _LIST_ITEM.match(line)
        if not match:
            continue
        text = match.group(1).strip()
        if not text:
            continue
        requirement_ids = _ID_PATTERN.findall(text)
        # Strip a "(category)" marker wherever it appears, since a requirement
        # reference often follows it.
        category = ""
        for marker in re.finditer(r"\(([^()]+)\)", text):
            candidate = marker.group(1).strip().casefold().replace(" ", "_")
            if candidate in _KNOWN_CATEGORIES:
                category = candidate
                text = (text[: marker.start()] + " " + text[marker.end() :]).strip()
                break
        title = _ID_PATTERN.sub("", text).strip(" -–—:;,")
        if not title:
            continue
        records.append(
            {
                "title": title,
                "description": "",
                "category": category or "functional",
                "requirement_ids": requirement_ids,
            }
        )
    return records


def read_markdown_scenarios(path: Path, text: str) -> list[dict[str, Any]]:
    """Read scenarios from a markdown or plain-text document.

    Recognises, in order: a pipe table with a title column, then a bulleted or
    numbered list. Unstructured prose raises a clear error rather than being
    guessed at - that input belongs to the `document` entry point.
    """
    lines = text.splitlines()
    records = _read_markdown_table(lines)
    if records:
        return records
    records = _read_list_items(lines)
    if records:
        return records
    msg = (
        f"No structured scenarios found in {path}. Expected a markdown table "
        "with a title column, or a bulleted/numbered list of scenarios. For "
        "free-form prose, run without --from so the requirement analyzer "
        "processes the document instead."
    )
    raise DocumentLoadError(msg)
